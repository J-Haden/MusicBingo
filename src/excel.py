from pathlib import Path
from openpyxl import load_workbook

def create_excel_card(card, title, template_path, output_path):
    """
    Populate an Excel bingo template.

    Parameters:
        card: 5x5 list containing song titles
        title: Game title
        template_path: Path to Excel template
        output_path: Where to save completed workbook
    """
    
    workbook = load_workbook(template_path)
    sheet = workbook.active
    
    sheet['A1'] = title
    
    start_row = 3
    start_col = 1
    
    for row_index, row in enumerate(card):
        for col_index, song in enumerate(row):
            cell = sheet.cell(
                row = start_row + row_index,
                column = start_col + col_index
            )
            
            cell.value = song
    
    workbook.save(output_path)